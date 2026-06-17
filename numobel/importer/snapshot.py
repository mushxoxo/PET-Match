"""Detect and restore NUMOBEL snapshot workbooks (the export round-trip format).

A *snapshot* is the self-contained ``.xlsx`` produced by
:func:`numobel.exporter.run_export.export`: one worksheet per database table,
plus embedded product photos. This module restores such a file verbatim
(replacing the current catalog) and exposes :func:`import_workbook`, which
auto-detects whether a given file is a snapshot or the original brand-sheet
master workbook and routes it to the right loader.

The format marker lives in a ``_meta`` sheet so detection never has to guess
from sheet names alone.
"""

from __future__ import annotations

import base64
import sqlite3

from openpyxl import load_workbook

from numobel import db
from numobel.importer.run_import import build
from numobel.sync import serialize
from numobel.sync.serialize import EXPORT_RESTORE_ORDER

#: Identifier written to (and matched in) the snapshot ``_meta`` sheet.
FORMAT = "numobel-snapshot"
#: Bump when the on-disk layout changes incompatibly.
SNAPSHOT_VERSION = 1

__all__ = [
    "FORMAT",
    "SNAPSHOT_VERSION",
    "EXPORT_RESTORE_ORDER",
    "is_snapshot",
    "restore",
    "import_workbook",
]


def is_snapshot(excel_path: str) -> bool:
    """Return True when ``excel_path`` is a NUMOBEL snapshot workbook.

    Reads only the ``_meta`` sheet; any load error (corrupt/locked/non-xlsx)
    is treated as "not a snapshot" so the caller falls back to ``build()``.
    """
    try:
        wb = load_workbook(excel_path, read_only=True)
    except Exception:  # noqa: BLE001 — unreadable means "not a snapshot"
        return False
    try:
        if "_meta" not in wb.sheetnames:
            return False
        for row in wb["_meta"].iter_rows(values_only=True):
            if row and row[0] == "format":
                return row[1] == FORMAT
        return False
    finally:
        wb.close()


def _read_sheet(wb, name: str):
    """Yield each data row of a snapshot table sheet as a ``{column: value}`` dict.

    Row 1 holds the column names; trailing ``None`` header padding is dropped and
    empty cells map to ``None`` (SQL ``NULL``).
    """
    if name not in wb.sheetnames:
        return
    rows = wb[name].iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    cols = [c for c in header if c is not None]
    for row in rows:
        yield {col: (row[i] if i < len(row) else None) for i, col in enumerate(cols)}


def _restore_images(wb, conn: sqlite3.Connection) -> int:
    """Recreate embedded photos as files and repoint ``products.image_path``.

    Photos are stored in the ``_images`` sheet as base64 split across ``seq``
    ordered chunks. Each is decoded, written under :func:`db.images_dir`, and the
    owning product's ``image_path`` is set to the portable ``images/<file>`` form
    used elsewhere. Products whose photo was not embedded have their (now broken)
    ``image_path`` cleared. Returns the number of photos restored.
    """
    if "_images" not in wb.sheetnames:
        # No image sheet at all: leave image_path values untouched.
        return 0

    chunks: dict[tuple[int, str], dict[int, str]] = {}
    for rec in _read_sheet(wb, "_images"):
        pid, filename, chunk = rec.get("product_id"), rec.get("filename"), rec.get("chunk_b64")
        if pid is None or filename is None or chunk is None:
            continue
        chunks.setdefault((int(pid), str(filename)), {})[int(rec.get("seq") or 0)] = chunk

    images_dir = db.images_dir()
    if chunks:
        images_dir.mkdir(parents=True, exist_ok=True)

    restored: set[int] = set()
    for (pid, filename), parts in chunks.items():
        data = "".join(parts[s] for s in sorted(parts))
        (images_dir / filename).write_bytes(base64.b64decode(data))
        conn.execute(
            "UPDATE products SET image_path = ? WHERE id = ?",
            (f"images/{filename}", pid),
        )
        restored.add(pid)

    # Clear any image_path that points at a photo we did not carry along.
    if restored:
        placeholders = ",".join("?" * len(restored))
        conn.execute(
            f"UPDATE products SET image_path = NULL "
            f"WHERE image_path IS NOT NULL AND id NOT IN ({placeholders})",
            list(restored),
        )
    else:
        conn.execute(
            "UPDATE products SET image_path = NULL WHERE image_path IS NOT NULL"
        )
    return len(restored)


def restore(excel_path: str, conn: sqlite3.Connection) -> dict:
    """Restore a snapshot into ``conn``, replacing the existing catalog.

    Returns a summary dict shaped like :func:`numobel.importer.run_import.build`'s
    so the import UI can report results uniformly.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        db.create_schema(conn)
        db.reset_catalog(conn)

        # audit_log round-trips with the snapshot (a full DB dump) but is
        # machine-local for sync, so reset_catalog deliberately keeps it. When
        # the snapshot carries an audit_log sheet, replace local history with the
        # snapshot's (clear first so rows are not appended); an older snapshot
        # without the sheet leaves local history untouched.
        if "audit_log" in wb.sheetnames:
            conn.execute("DELETE FROM audit_log")

        link_tallies = {"resolved": 0, "unresolved": 0, "external": 0}
        product_count = 0
        price_count = 0
        audit_count = 0

        for table in EXPORT_RESTORE_ORDER:
            records = list(_read_sheet(wb, table))
            if not records:
                continue
            columns = list(records[0].keys())
            rows = [[rec[c] for c in columns] for rec in records]
            serialize.restore_table(conn, table, columns, rows)
            if table == "products":
                product_count += len(records)
            elif table == "prices":
                price_count += len(records)
            elif table == "audit_log":
                audit_count += len(records)
            elif table == "color_links":
                for rec in records:
                    status = rec.get("status")
                    if status in link_tallies:
                        link_tallies[status] += 1

        images = _restore_images(wb, conn)
        conn.commit()
    finally:
        wb.close()

    return {
        "total_products": product_count,
        "prices": price_count,
        "links": link_tallies,
        "total_links": sum(link_tallies.values()),
        "audit_log": audit_count,
        "images": images,
    }


def import_workbook(excel_path: str, conn: sqlite3.Connection) -> dict:
    """Load ``excel_path`` into ``conn``, auto-detecting the workbook format.

    Snapshot files are restored verbatim; anything else is treated as the
    original brand-sheet master workbook and handed to ``build()``. Both paths
    replace the current catalog and return a compatible summary dict.
    """
    if is_snapshot(excel_path):
        return restore(excel_path, conn)
    return build(excel_path=excel_path, conn=conn)
