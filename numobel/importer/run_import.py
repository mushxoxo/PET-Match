"""Two-pass Excel import pipeline for the NUMOBEL catalog.

``build()`` performs a fresh build: resets the catalog, inserts brands, loads
products from every brand sheet (pass 1), synthesizes Numobel hub products from
the distinct Numobel codes referenced across all mapping cells, then resolves
cross-reference links (pass 2) and inserts the price table.

Run ``python -m numobel.importer.run_import`` to build against the real
workbook and print a summary.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .. import db
from . import sheets
from .refparse import parse_ref

DEFAULT_EXCEL_PATH = str(
    Path(__file__).resolve().parent.parent.parent
    / "my_excel"
    / "NUMOBEL_ACOUSTICS_COLOR_MAPS.xlsx"
)

_DIGITS_RE = re.compile(r"\d+")

#: Sentinel marking a shade-number that maps to more than one product within a
#: brand — such keys are never used for resolution (to avoid wrong matches).
_AMBIGUOUS = object()


def _sku_number(sku: str | None) -> str | None:
    """Return a SKU's first digit-run with leading zeros stripped, or None.

    References and stored SKUs sometimes agree on brand and shade number but
    differ in formatting — zero-padding (``AT05`` vs ``AT5``) or brand-prefix
    variant (``BOL06`` vs ``B6``, ``UT19`` vs ``UTAB19``). Since the canonical
    brand is already known from the resolved brand id, matching on the shade
    number alone bridges those formats: ``_sku_number('BOL06') == '6'``.
    """
    if not sku:
        return None
    m = _DIGITS_RE.search(sku)
    if not m:
        return None
    return str(int(m.group()))


def _load_rows(wb, name):
    """Return a list of value-only row tuples for a sheet."""
    return list(wb[name].iter_rows(values_only=True))


def build(
    db_path: str = db.DEFAULT_DB_PATH,
    excel_path: str = DEFAULT_EXCEL_PATH,
    conn=None,
) -> dict:
    """Build the catalog from a workbook.

    When ``conn`` is given (e.g. the running app's live connection) it is used
    and left open, so the caller sees the new rows immediately and there is no
    second connection competing for the database lock. Otherwise a connection
    is opened from ``db_path`` and closed before returning.
    """
    own_conn = conn is None
    if own_conn:
        conn = db.connect(db_path)

    try:
        return _build(conn, excel_path)
    finally:
        if own_conn:
            conn.close()


def _build(conn, excel_path: str) -> dict:
    db.create_schema(conn)
    db.reset_catalog(conn)

    created_at = datetime.now().isoformat()
    wb = load_workbook(excel_path, data_only=True)

    # --- Insert brands ---------------------------------------------------
    brand_id: dict[str, int] = {}
    brand_has_sheet: dict[str, int] = {}
    for code, name, has_sheet in sheets.BRANDS:
        cur = conn.execute(
            "INSERT INTO brands(code, name, has_sheet) VALUES(?,?,?)",
            (code, name, has_sheet),
        )
        brand_id[code] = cur.lastrowid
        brand_has_sheet[code] = has_sheet

    # --- PASS 1: products -----------------------------------------------
    # index: (brand_id, normalized_code) -> product_id
    index: dict[tuple[int, str], int] = {}
    # num_index: (brand_id, shade_number) -> product_id | _AMBIGUOUS — a
    # format-insensitive fallback used in pass 2 when an exact SKU match fails.
    num_index: dict[tuple[int, str], object] = {}
    per_sheet_counts: dict[str, int] = {}
    # Keep the raw products around so pass 2 can re-walk their mapping cells.
    all_raw: list[sheets.RawProduct] = []
    # Track distinct Numobel codes seen across all mapping cells.
    numobel_codes: dict[str, str | None] = {}  # normalized -> color_name

    for sheet_name, spec in sheets.BRAND_REGISTRY.items():
        rows = _load_rows(wb, sheet_name)
        raws = spec.extract(rows)
        count = 0
        for rp in raws:
            bid = brand_id[rp.brand_code]
            # Source sheets contain a few duplicate SKUs (e.g. PNV ACP42,
            # Tranquil TRIR38). UNIQUE(brand_id, sku) forbids them; keep the
            # first occurrence and skip later duplicates.
            if rp.sku and (bid, rp.sku) in index:
                continue
            extra_json = json.dumps(rp.extra) if rp.extra else None
            cur = conn.execute(
                """INSERT INTO products
                   (brand_id, sku, shade_no, color_name, thickness, self_label,
                    category, extra_json, source_sheet, source_row)
                   VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (bid, rp.sku, rp.shade_no, rp.color_name, rp.thickness,
                 rp.self_label, rp.category, extra_json,
                 rp.source_sheet, rp.source_row),
            )
            pid = cur.lastrowid
            if rp.sku:
                index[(bid, rp.sku)] = pid
                num = _sku_number(rp.sku)
                if num is not None:
                    nkey = (bid, num)
                    # Two products in one brand sharing a shade number make the
                    # number-only key ambiguous; mark it so it is never matched.
                    num_index[nkey] = _AMBIGUOUS if nkey in num_index else pid
            rp._pid = pid  # type: ignore[attr-defined]
            all_raw.append(rp)
            count += 1

            # Collect Numobel codes referenced by this row's mapping cells.
            for cell in rp.mappings:
                ref = parse_ref(cell.value, default_brand=cell.default_brand)
                if ref is None:
                    continue
                if sheets.resolve_brand(ref.brand_code) == sheets.HUB_BRAND:
                    numobel_codes.setdefault(ref.normalized, ref.name)
        per_sheet_counts[sheet_name] = count

    # --- Synthesize Numobel hub products --------------------------------
    nu_bid = brand_id[sheets.HUB_BRAND]
    nu_count = 0
    for norm, color in sorted(numobel_codes.items()):
        cur = conn.execute(
            """INSERT INTO products
               (brand_id, sku, color_name, self_label, source_sheet)
               VALUES(?,?,?,?,?)""",
            (nu_bid, norm, color, norm, "(synthesized)"),
        )
        index[(nu_bid, norm)] = cur.lastrowid
        nu_count += 1
    per_sheet_counts["NUMOBEL(synth)"] = nu_count

    # --- PASS 2: links ---------------------------------------------------
    tallies = {"resolved": 0, "unresolved": 0, "external": 0}
    seen_links: set[tuple[int, object, str]] = set()

    for rp in all_raw:
        from_pid = rp._pid  # type: ignore[attr-defined]
        for cell in rp.mappings:
            ref = parse_ref(cell.value, default_brand=cell.default_brand)
            if ref is None:
                continue
            canon = sheets.resolve_brand(ref.brand_code)
            if canon is None:
                continue  # unknown brand prefix; skip

            to_pid = None
            if canon in sheets.EXTERNAL_BRANDS:
                status = "external"
            else:
                # sheet brand or NUMOBEL: try to resolve via index
                tbid = brand_id.get(canon)
                to_pid = index.get((tbid, ref.normalized)) if tbid else None
                # Fallback: match on shade number alone so refs that differ
                # only by zero-padding or brand-prefix variant still resolve
                # (e.g. 'AT05'->'AT5', 'BOL06'->'B6', 'UT19'->'UTAB19').
                # NUMOBEL is excluded — its hub codes reuse numbers across
                # prefixes (NW.../NU...), so number-only matching is ambiguous.
                if to_pid is None and tbid and canon != sheets.HUB_BRAND:
                    num = _sku_number(ref.normalized)
                    cand = num_index.get((tbid, num)) if num is not None else None
                    if cand is not None and cand is not _AMBIGUOUS:
                        to_pid = cand
                status = "resolved" if to_pid is not None else "unresolved"

            # Dedupe identical links (same from/to_product/normalized).
            key = (from_pid, to_pid, ref.normalized)
            if key in seen_links:
                continue
            seen_links.add(key)

            conn.execute(
                """INSERT INTO color_links
                   (from_product_id, to_product_id, to_brand_code, raw_ref,
                    normalized, status, source, created_at)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (from_pid, to_pid, canon, ref.raw, ref.normalized,
                 status, "import", created_at),
            )
            tallies[status] += 1

    # --- Prices ----------------------------------------------------------
    price_rows = sheets.extract_prices(_load_rows(wb, sheets.PRICE_SHEET))
    for pr in price_rows:
        conn.execute(
            """INSERT INTO prices
               (seller, mrp, mrp_sft, dp, dp_sft, profit, discount,
                cust_price, cust_price_sft)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            pr,
        )

    conn.commit()

    # NOTE: resolved cross-reference links are folded into transitive color
    # families by db.migrate(), which the app runs at startup (and tests run
    # explicitly). Keeping build() free of that step leaves the resolved links
    # inspectable here and in the import summary.

    total_products = sum(per_sheet_counts.values())
    summary = {
        "products_by_sheet": per_sheet_counts,
        "total_products": total_products,
        "links": tallies,
        "total_links": sum(tallies.values()),
        "prices": len(price_rows),
    }
    return summary


def _print_summary(summary: dict) -> None:
    print("=== NUMOBEL import summary ===")
    print("Products by sheet:")
    for sheet, n in summary["products_by_sheet"].items():
        print(f"  {sheet:18s} {n}")
    print(f"Total products: {summary['total_products']}")
    print("Link tallies:")
    for status, n in summary["links"].items():
        print(f"  {status:12s} {n}")
    print(f"Total links: {summary['total_links']}")
    print(f"Prices: {summary['prices']}")


if __name__ == "__main__":
    _print_summary(build())
