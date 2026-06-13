"""Export the live catalog to a single, self-contained ``.xlsx`` snapshot.

The snapshot is a full database dump — one worksheet per table, plus embedded
product photos — that :mod:`numobel.importer.snapshot` can restore verbatim so a
colleague can import it and continue working. It is deliberately distinct from
the original brand-sheet master workbook read by
:func:`numobel.importer.run_import.build`; that format cannot represent
user-added products, edited prices, attached photos, or folded color groups.
"""

from __future__ import annotations

import base64
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from numobel import db
from numobel.importer.snapshot import FORMAT, SNAPSHOT_VERSION

#: Tables dumped, in foreign-key-safe order (consumed in the same order on
#: restore). Machine-local tables (``audit_log``, ``settings``) are excluded.
SNAPSHOT_TABLES = ("brands", "color_groups", "products", "color_links", "prices")

#: Base64 chunk size, kept under Excel's ~32,767-character per-cell limit.
_CHUNK = 32000


def _dump_table(wb: Workbook, conn: sqlite3.Connection, table: str) -> int:
    """Write one table to its own sheet (header row = column names). Returns rows."""
    ws = wb.create_sheet(table)
    cur = conn.execute(f"SELECT * FROM {table}")
    cols = [d[0] for d in cur.description]
    ws.append(cols)
    n = 0
    for row in cur:
        ws.append([row[c] for c in cols])
        n += 1
    return n


def _resolve_image(path: str) -> Path:
    """Resolve a stored ``image_path`` to an absolute path (matches detail_panel)."""
    p = Path(path)
    return p if p.is_absolute() else db.base_dir() / p


def _embed_images(wb: Workbook, conn: sqlite3.Connection) -> int:
    """Embed each product photo as base64 chunks in ``_images``. Returns count."""
    ws = wb.create_sheet("_images")
    ws.append(["product_id", "filename", "seq", "chunk_b64"])
    count = 0
    rows = conn.execute(
        "SELECT id, image_path FROM products "
        "WHERE image_path IS NOT NULL AND image_path <> ''"
    ).fetchall()
    for r in rows:
        src = _resolve_image(r["image_path"])
        if not src.is_file():
            continue  # broken reference on this machine — nothing to carry
        encoded = base64.b64encode(src.read_bytes()).decode("ascii")
        # The on-disk name is already unique per product (detail_panel stores
        # it as "<product_id>_<basename>"), so reuse it verbatim — keeps the
        # path stable across repeated export/import cycles.
        filename = src.name
        for seq, start in enumerate(range(0, len(encoded), _CHUNK)):
            ws.append([r["id"], filename, seq, encoded[start : start + _CHUNK]])
        count += 1
    return count


def export(excel_path: str, conn: sqlite3.Connection) -> dict:
    """Write a full-catalog snapshot to ``excel_path``. Returns a summary dict."""
    wb = Workbook()
    meta = wb.active  # reuse the default sheet as _meta
    meta.title = "_meta"
    meta.append(["key", "value"])
    meta.append(["format", FORMAT])
    meta.append(["version", SNAPSHOT_VERSION])
    meta.append(["exported_at", datetime.now().isoformat(timespec="seconds")])

    counts = {t: _dump_table(wb, conn, t) for t in SNAPSHOT_TABLES}
    images = _embed_images(wb, conn)
    wb.save(excel_path)

    return {
        "total_products": counts["products"],
        "prices": counts["prices"],
        "links": counts["color_links"],
        "images": images,
    }
